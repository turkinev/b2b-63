from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os, requests as req

app = Flask(__name__)
CORS(app)

MM_HOOK          = 'https://mm.63pokupki.ru:8443/hooks/k7xh9osx5tr1pbyr1b6y3rqcba'
MM_HOOK_PVZ      = 'https://mm.63pokupki.ru:8443/hooks/ycpetuzfn78u881yx793cfghpy'
MM_HOOK_SUPPLIER = 'https://mm.63pokupki.ru:8443/hooks/sds81m1gyjnw7guk8xjy9agqcw'

def notify_mm_hook(hook, text):
    try:
        req.post(hook, json={'text': text}, timeout=5)
    except:
        pass

FILE        = 'заявки.xlsx'
VISITS_FILE = 'посетители.xlsx'

def get_wb(path, headers):
    if os.path.exists(path):
        return load_workbook(path)
    wb = Workbook()
    wb.active.append(headers)
    return wb

@app.route('/submit', methods=['POST'])
def submit():
    d = request.get_json()
    wb = get_wb(FILE, ['Дата', 'Компания', 'Контакт', 'ИНН', 'Комментарий'])
    wb.active.append([datetime.now().strftime('%d.%m.%Y %H:%M'), d['company'], d['contact'], d.get('inn', ''), d.get('product', '')])
    wb.save(FILE)
    notify_mm_hook(MM_HOOK, f"📥 **Новая B2B заявка**\n**ИНН:** {d.get('inn','—')}\n**Компания:** {d['company']}\n**Контакт:** {d['contact']}\n**Комментарий:** {d.get('product','—')}")
    return jsonify({'ok': True})

@app.route('/track', methods=['POST'])
def track():
    d = request.get_json()
    ip = request.headers.get('X-Forwarded-For', request.remote_addr).split(',')[0].strip()

    country, city = '', ''
    try:
        geo = req.get(f'http://ip-api.com/json/{ip}?lang=ru&fields=country,city', timeout=2).json()
        country = geo.get('country', '')
        city    = geo.get('city', '')
    except:
        pass

    ua = request.headers.get('User-Agent', '')
    row = [
        datetime.now().strftime('%d.%m.%Y %H:%M:%S'),
        ip, country, city, ua,
        d.get('referrer', ''),
        d.get('screen', ''),
        d.get('language', ''),
        d.get('user_id', ''),
    ]
    headers = ['Дата/Время', 'IP', 'Страна', 'Город', 'Браузер (User-Agent)', 'Источник (Referrer)', 'Экран', 'Язык', 'User ID']
    wb = get_wb(VISITS_FILE, headers)
    wb.active.append(row)
    wb.save(VISITS_FILE)
    return jsonify({'ok': True})

@app.route('/submit-pvz', methods=['POST'])
def submit_pvz():
    d = request.get_json()
    wb = get_wb('заявки_пвз.xlsx', ['Дата', 'ИНН', 'Компания', 'Адреса', 'Кол-во ПВЗ', 'Контакт', 'Комментарий'])
    wb.active.append([datetime.now().strftime('%d.%m.%Y %H:%M'), d.get('inn',''), d['company'], d.get('address',''), d.get('count',''), d['contact'], d.get('comment','')])
    wb.save('заявки_пвз.xlsx')
    notify_mm_hook(MM_HOOK_PVZ, f"🏪 **Новая заявка ПВЗ**\n**ИНН:** {d.get('inn','—')}\n**Компания:** {d['company']}\n**Адреса:** {d.get('address','—')}\n**Кол-во ПВЗ:** {d.get('count','—')}\n**Контакт:** {d['contact']}\n**Комментарий:** {d.get('comment','—')}")
    return jsonify({'ok': True})

@app.route('/submit-supplier', methods=['POST'])
def submit_supplier():
    d = request.get_json()
    wb = get_wb('заявки_поставщики.xlsx', ['Дата', 'ИНН', 'Компания', 'Категория', 'Сайт', 'Контакт', 'Комментарий'])
    wb.active.append([datetime.now().strftime('%d.%m.%Y %H:%M'), d.get('inn',''), d['company'], d.get('category',''), d.get('site',''), d['contact'], d.get('comment','')])
    wb.save('заявки_поставщики.xlsx')
    notify_mm_hook(MM_HOOK_SUPPLIER, f"🚚 **Новая заявка поставщика**\n**ИНН:** {d.get('inn','—')}\n**Компания:** {d['company']}\n**Категория:** {d.get('category','—')}\n**Сайт:** {d.get('site','—')}\n**Контакт:** {d['contact']}\n**Комментарий:** {d.get('comment','—')}")
    return jsonify({'ok': True})

@app.route('/download/secret123/заявки')
def download_orders():
    return send_file(FILE, as_attachment=True)

@app.route('/download/secret123/посетители')
def download_visits():
    return send_file(VISITS_FILE, as_attachment=True)

@app.route('/download/secret123/пвз')
def download_pvz():
    return send_file('заявки_пвз.xlsx', as_attachment=True)

@app.route('/download/secret123/поставщики')
def download_suppliers():
    return send_file('заявки_поставщики.xlsx', as_attachment=True)

app.run(port=5000)
